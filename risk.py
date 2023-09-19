from random import *
from datetime import datetime,timedelta
from datetime import date
from dateutil.relativedelta import relativedelta
from faker import Faker
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from faker.providers import DynamicProvider
import random
fake=Faker()
from openpyxl import Workbook
from openpyxl.styles import Font
wb=Workbook()
ws=wb.active

ws.column_dimensions['A'].width=20
def calculate_score(age):
    if  age<=35:
        return fake.random_element(elements=('0.9','1'))
    elif age>=36 and age<=45:
        return fake.random_element(elements=('0.8','0.7'))
    elif age>=46 and age<=55:
        return fake.random_element(elements=('0.5','0.6'))
    elif age>=56 and age<=65:
        return  0.4
    else:
        return fake.random_element(elements=('0.2','0.3'))
def calculate_value(value): 
    if age<=35:
        return "excellent"
    elif age>=36 and age<=45:
        return "V.good"
    elif age>=46 and age<=55:
        return "good"
    elif age>=56 and age<=65:
        return "avg"
    else:
        return "bad" 
def calculate_Income_score(Income):
    if Income>=200000:
        return fake.random_element(elements=('0.9','1'))
    elif Income>=150000 and Income<199999:
        return fake.random_element(elements=('0.8','0.7'))
    elif Income>=100000 and Income<149000:
        return fake.random_element(elements=('0.5','0.6'))
    elif Income>=60000 and Income<100000:
        return fake.random_element(elements=('0.3','0.7'))
    else:
        return 0 
     
def calculate_Income_value(value):
    if Income>=200000:
        return "Excellent"
    elif Income>=150000 and Income<199999:
        return "V.Good"
    elif Income>=100000 and Income<149000:
        return "Good"
    elif Income>=60000 and Income<100000:
        return "Avg"
    else:
        return "bad"      
def calculate_smoking_status(smoking_status):
    if smoking_status=="smoker":
        return 0.5
    else:
        return fake.random_element(elements=('0.8','0.7','0.9','0.6'))
def calculate_smoking_status_value(value):
     if smoking_status=="smoker":
         return "high risk"
     else:
         return fake.random_element(elements=('excellent','v.good','low risk'))         
def calculate_occupation_risk(occpation):
    if occpation=="Policeman":
        return 0.4
    elif occpation=="Electrician":
        return 0.5 
    else:
        return 0.9 
def calculate_occupation_risk_value(occpation):
    if occpation=="Policeman":
        return "high risk"
    elif occpation=="Electrician":
        return "medium risk" 
    else:
        return "low risk"  
def calculate_medical_history(family_medical_history):
     if family_medical_history=="significant history":
         return 0.4
     else:
         return 0.9
def calculate_medical_history_value(value):
     if family_medical_history=="significant history":
         return "high risk"
     else:
         return "low risk"
def calculate_MVR(mvr):               
    if mvr=="clean":
        return 0.8
    elif mvr=="Minor Violations" :
        return 0.5
    else:
        return 0.4
def calculate_MVR_value(mvr):               
    if mvr=="clean":
        return "low risk"
    elif mvr=="Minor Violations" :
        return "medium risk"
    else:
        return "high risk"
def calculate_MIB(mib):
    if mib=="negative report":
        return 0.5
    else:
        return 0.6
def calculate_MIB_value(mib):
    if mib=="negative report":
        return "high risk"
    else:
        return "low risk"
def calculate_mib_score(mib):
    if value=="excellent":
        return 0.9
    elif value=="good":
        return 0.6
    else:
        return 0.4 
def calculate_credit(value):
    if value=="average" :
        return 0.4
    elif value=="Excellent":
        return 0.9
    else:
        return 0.6     
def calculate_height(height,weight):
    if height>="160" and height<="170":
        if weight>="60" and weight<="70":
            return "low risk" 
        else:
            return "high risk"
    elif height>="171" and height<="180":
        if weight>="70" and weight<="85":
            return "low risk" 
        else:
            return "high risk"
      
def calculate_height_value(height,weight):
    if height>="160" and height<="170":
        if weight>="60" and weight<="70":
            return "0.8" 
        else:
            return "0.5"
    elif height>="171" and height<="180":
        if weight>="70" and weight<="85":
            return "0.8" 
        else:
            return "0.5" 
def under_writing(under_writing_decision) :
    if under_writing_decision=="Declined":
        return "Declined"
    elif under_writing_decision=="Accepted":
        return fake.random_element(elements=('Lapsed','inforce','claim settled'))
    else:
        return "unknown"
def policy_status_claim(policy_effective_date,First_Unpaid_premium_date,policy_status):
    if policy_status=="claim settled":
        
        # Generate a random year, month, and day within the range
        
        return random_date
def policy_status_claim_settle(policy_status,random_date):
    if policy_status == "claim settled":
        
        # random_date = policy_status_claim(policy_effective_date,First_Unpaid_premium_date,policy_status)
        Date_of_settlement =  random_date + relativedelta(months=1)
        
        return Date_of_settlement

# Call the second function with the generated random_date
    
         
def policy_status_claim_cause(policy_status):
    if policy_status=="claim settled":
       
        
        return fake.random_element(elements=('Accident','Accident'))                
ws['A1']='sl.no'
ws['B1']=''
ws['c1']='details'
for i in range(1,11):
    heading=f"ph{i}"
    cell=ws.cell(row=1,column=i+3)
    cell.value=heading
for cell in ws[1]:
  cell.font=Font(bold=True) 
  cell.alignment=Alignment(horizontal="center") 
ws.append([""])
ws.merge_cells('A3:A7')
merged_cell=ws['A3']
merged_cell.value="1"
merged_cell.alignment=Alignment(horizontal='center',vertical='center')
font=Font(bold=True)
merged_cell.font=font

ws.merge_cells('B3:B7')
merged_cell1=ws['B3']
merged_cell1.value="Policy"
merged_cell1.alignment=Alignment(horizontal='center',vertical='center')
merged_cell1.font=font


ws['C3'].value="Policy Number"

ws['C4'].value="Policy Status"

ws['C6'].value="Policy Effective/Declined Date "

ws['C7'].value="Policy Expiry Date"

ws.merge_cells('A9:A13')
merged_cell=ws['A9']
merged_cell.value="2"
merged_cell.alignment=Alignment(horizontal='center',vertical='center')
font=Font(bold=True)
merged_cell.font=font

ws.merge_cells('B9:B13')
merged_cell1=ws['B9']
merged_cell1.value="Life Assured"
merged_cell1.alignment=Alignment(horizontal='center',vertical='center')
merged_cell1.font=font


ws['C9'].value="Surname"

ws['C10'].value="Given Name"

ws['C11'].value="Birth Date"

ws['C12'].value="Gender"

ws['C13'].value="Occupation"

ws.merge_cells('A15:A21')
merged_cell=ws['A15']
merged_cell.value="3"
merged_cell.alignment=Alignment(horizontal='center',vertical='center')
font=Font(bold=True)
merged_cell.font=font

ws.merge_cells('B15:B21')
merged_cell1=ws['B15']
merged_cell1.value="Address"
merged_cell1.alignment=Alignment(horizontal='center',vertical='center')
merged_cell1.font=font
values=[
    "address type","address","city","state province","postal code","country"
]
column='c'
start_row=16
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value

ws.merge_cells('A24:A30')
merged_cell=ws['A24']
merged_cell.value="4"
merged_cell.alignment=Alignment(horizontal='center',vertical='center')
font=Font(bold=True)
merged_cell.font=font

ws.merge_cells('B24:B30')
merged_cell1=ws['B24']
merged_cell1.value="Product"
merged_cell1.alignment=Alignment(horizontal='center',vertical='center')
merged_cell1.font=font
values=[
    "product type","product name","Face Amt","Premium Amt","Premium Frequency","First_Unpaid_premium_date"
]
column='c'
start_row=25
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value
ws.merge_cells('A32:A35')
merged_cell=ws['A32']
merged_cell.value="5"
merged_cell.alignment=Alignment(horizontal='center',vertical='center')
font=Font(bold=True)
merged_cell.font=font

ws.merge_cells('B32:B35')
merged_cell1=ws['B32']
merged_cell1.value="carrier"
merged_cell1.alignment=Alignment(horizontal='center',vertical='center') 
merged_cell1.font=font
values=[
    "carrier name","carrier code"
]
column='c'
start_row=33
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value
ws.merge_cells('A37:A39')
merged_cell=ws['A37']
merged_cell.value="6"
merged_cell.alignment=Alignment(horizontal='center',vertical='center')
font=Font(bold=True)
merged_cell.font=font


ws.merge_cells('B37:B39')
merged_cell1=ws['B37']
merged_cell1.value="agent"
merged_cell1.alignment=Alignment(horizontal='center',vertical='center') 
merged_cell1.font=font
values=[
    "agent name","agent code","Commission Pct"
]
column='c'
start_row=37
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value  
ws.merge_cells('A41:A44')
merged_cell=ws['A41']
merged_cell.value="7"
merged_cell.alignment=Alignment(horizontal='center',vertical='center')
font=Font(bold=True)
merged_cell.font=font

ws.merge_cells('B41:B44')
merged_cell1=ws['B41']
merged_cell1.value="Underwriting"
merged_cell1.alignment=Alignment(horizontal='center',vertical='center') 
merged_cell1.font=font
values=[
    "Underwriting Decision Cd"," Underwriting Decision "
]
column='c'
start_row=42
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value     
ws.merge_cells('A46:A89')
merged_cell=ws['A46']
merged_cell.value="8"
merged_cell.alignment=Alignment(horizontal='center',vertical='center')
font=Font(bold=True)
merged_cell.font=font

ws.merge_cells('B46:B89')
merged_cell1=ws['B46']
merged_cell1.value="Underwriting Parameters"
merged_cell1.alignment=Alignment(horizontal='center',vertical='center')
merged_cell1.font=font 
values=[
    "age","value","score"
]
column='c'
start_row=50
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value 
values=[
    "Income","value","score"
]
column='c'
start_row=46
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value 
values=[
    "smokingstatus","value","score"
]
column='c'
start_row=54
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value 
values=[
    "occupation risk","value","score"
]
column='c'
start_row=59
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value 
values=[
    "family_medical_history","value","score"
]
column='c'
start_row=64
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value 
values=[
    "bmi","height","weight","value","score"
]
column='c'
start_row=68
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value 
values=[
    "mvr","value","score"
]
column='c'
start_row=74
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value 
values=[
    "mib","value","score"
]
column='c'
start_row=79
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value    
values=[
    "credit_rating","value","score"
]
column='c'
start_row=83
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value 
values=[
    "social_media","value","score"
]
column='c'
start_row=87
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value
ws.merge_cells('A41:A44')
merged_cell=ws['A41']
merged_cell.value="7"
merged_cell.alignment=Alignment(horizontal='center',vertical='center')
font=Font(bold=True)
merged_cell.font=font

ws.merge_cells('B41:B44')
merged_cell1=ws['B41']
merged_cell1.value="Underwriting"
merged_cell1.alignment=Alignment(horizontal='center',vertical='center') 
merged_cell1.font=font
values=[
    " Underwriting Decision "
]
column='c'
start_row=43
for index,value in enumerate(values):
    ws[column+str(start_row+index)].value=value     
ws.merge_cells('A91:A96')
merged_cell=ws['A91']
merged_cell.value="9"
merged_cell.alignment=Alignment(horizontal='center',vertical='center')
font=Font(bold=True)
merged_cell.font=font

ws.merge_cells('B91:B96')
merged_cell1=ws['B91']
merged_cell1.value="Adverse Medical Details"
merged_cell1.alignment=Alignment(horizontal='center',vertical='center')
merged_cell1.font=font 
ws['C91'].value="name"

ws['C92'].value="discription"

ws['C95'].value="name"

ws['C96'].value="discription"

ws.merge_cells('B98:B101')
merged_cell1=ws['B98']
merged_cell1.value="Claim Details"
merged_cell1.alignment=Alignment(horizontal='center',vertical='center')
merged_cell1.font=font

ws.merge_cells('A98:A101')
merged_cell=ws['A98']
merged_cell.value="10"
merged_cell.alignment=Alignment(horizontal='center',vertical='center')
font=Font(bold=True)
merged_cell.font=font
ws['C99'].value="Date of Death"

ws['C100'].value="Date of Claim settlement"

ws['C101'].value="Cause of Death "



for i in range(4,14):  # Generating 100 records
    print(i) 
    for row in ws.iter_rows(min_row=3,max_row=ws.max_row):
        for cell in row:
            cell.alignment=Alignment(horizontal="center")
    policy_number = fake.random_int(min=10000, max=99999)
    under_writing_decision=fake.random_element(elements=('Accepted','Accepted'))
    
    
    start_year = 1980
    end_year=2022
   
    random_year = random.randint(start_year, end_year)

# Generate random month and day
    month = random.randint(1, 12)
    day = random.randint(1, 28)

# Create the policy_effective_date
    policy_effective_date = date(random_year, month, day)
    

    policy_expiry_date=policy_effective_date.replace(year=random_year+20) 
    Surname=fake.name()
    given_name=fake.name()
    min_birth_year = random_year - 70  # 60 years before the effective date
    max_birth_year = random_year - 20   # 20 years before the effective date

# Generate a random date of birth within the specified range
    random_birth_year = random.randint(min_birth_year, max_birth_year)
    random_birth_month = random.randint(1, 12)
    random_birth_day = random.randint(1, 28)
    birthdate = date(random_birth_year, random_birth_month, random_birth_day)

    # birthdate = fake.date_of_birth(minimum_age=18, maximum_age=70) 
    Gender=fake.random_element(elements=('M','F'))
    occupation=fake.random_element(elements=('Engineer','Teacher','Doctor','Beautician','Interior Decorator','Lawyer','Electrician','Policeman','CSR','Researcher'))
    address_type=fake.random_element(elements=('residential','non_residential'))
    address=fake.random_element(elements=('3543 Upland Avenue','1715 Ashmor Drive',' 3656 Neuport Lane','1283 Bobcat Drive',' 1939 Longview Avenue','2815 Burke Street',' 728 Pooh Bear Lane',' 4719 Farland Avenue','1373 Oliver Street','4586 Ritter Street','881 Gladwell Street','2610 New Street'))
    city=fake.city()
    state_province=fake.state_abbr()
    postal_code=fake.zipcode()
    country=fake.random_element(elements=('US','US','US','US','US'))
    product_type=fake.random_element(elements=('Term ','Term ','Term ','Term '))
    product_name=fake.random_element(elements=('Term Life Insurance - 20 ','Term Life Insurance - 20 '))
    Face_Amt=fake.random_int(min=100000,max=1000000)
    Premium_Amt=fake.random_int(100,999)
    Premium_Frequency=fake.random_element(elements=('yearly','halferly','quaterly')) 
    if Premium_Frequency == "yearly":
        First_Unpaid_premium_date= policy_effective_date.replace(year=policy_effective_date.year + 1)
    
    elif Premium_Frequency == "quaterly":
         First_Unpaid_premium_date= policy_effective_date+relativedelta(months=3)
    elif Premium_Frequency == "halferly":
         First_Unpaid_premium_date= policy_effective_date+relativedelta(months=6)     
    
    #      next_quarter_year = policy_effective_date.year + (month + 2) // 12
    #      next_quarter_date = datetime.datetime(next_quarter_year, next_quarter_month, 1)
    # print("Policy Effective Date:", policy_effective_date.strftime("%Y-%m-%d"))
    # print("First Unpaid Premium Date (Quarterly):", next_quarter_date.strftime("%Y-%m-%d"))
    # elif Premium_Frequency == "halfly":
    # # Calculate the next semi-annual date
    #      month = policy_effective_date.month
    #      next_semi_annual_month = (month + 5) % 12 + 1
    #      next_semi_annual_year = policy_effective_date.year + (month + 5) // 12

    else:
        First_Unpaid_premium_date=0
    # print("Premium frequency other than yearly or quarterly is not supported in this code.")

             
    carrier_name=fake.random_element(elements=("The Guardian Life Insurance Company of America","The Guardian Life Insurance Company of America")) 
    carrier_code=fake.random_element(elements=('abc123','def456','ghi789','jkl012','mn0345','pqr678','stu987','vwx908')) 
    Agency_name=fake.random_element(elements=('American Family Insurance','Lincoln National Corporation ','General Reinsurance Corporation ',' Willis Towers Watson PLC   ',' AmTrust Financial Services ','USI Insurance Services LLC   ')) 
    Agent_code=fake.random_element(elements=('abc123','def456','ghi789','jkl012','mn0345','pqr678','stu987','vwx908'))
    Commission_Pct=fake.random_element(elements=('10.5','10.5','10.5')) 
    
    # print("Policy status is not 'claim settled'.")

    
    
    policy_status=str(under_writing(under_writing_decision))     
    Income=(random.randint(1000,3000)*100)
    
    # date_of_birth = datetime.strptime(Birth_date_str, '%Y-%m-%d')
    # effective_date = datetime.strptime(policy_effective_date_str, '%Y-%m-%d')
    

    age = policy_effective_date.year - birthdate.year - ((policy_effective_date.month, policy_effective_date.day) < (birthdate.month, birthdate.day))
# Calculate the age
    # age = (effective_date - date_of_birth).days // 365

    # age=(random.randint(18,100))
    
    smoking_status=fake.random_element(elements=('smoker','non-smoker')) 
    # occpation_risk=fake.random_element(elements=('Engineer','Teacher','Doctor','Beautician','Interior Decorator','Lawyer','Electrician','Policeman','CSR','Researcher'))
    family_medical_history=fake.random_element(elements=('no significant history','significant history')) 
    height=fake.random_element(elements=('160','172','164','168'))  
    weight=fake.random_element(elements=('83','74','60','65')) 
     
      
    mvr=fake.random_element(elements=('clean','Minor Violations','Major Violations'))
    mib=fake.random_element(elements=('positive report','negative report')) 
    value=fake.random_element(elements=('average','Excellent','good')) 
    # First_Unpaid_premium_date=fake.random_element(elements=('average','Excellent','good'))
    # social_media=fake.random_element(elements=('presence','non presence')) 
    
    score=fake.random_element(elements=('0.9','0.6','0.85')) 
    fake.random_element(elements=('presence','non presence')) 
    name=fake.random_element(elements=('heart disease','heart disease')) 
    description=fake.random_element(elements=('The applicant has a history of heart disease','The applicant has a history of heart disease'))  
    name=fake.random_element(elements=('cancer','cancer'))  
    description=fake.random_element(elements=('The applicant has a history of cancer disease','The applicant has a history of cancer disease'))
    start_date =policy_effective_date
    end_date =First_Unpaid_premium_date

    random_year = random.randint(start_date.year, end_date.year)
    random_month = random.randint(1, 12)  # 1 to 12 represents January to December
    random_day = random.randint(1, 28)    # 1 to 31 represents possible days
    random_date = date(random_year, random_month, random_day)  
    cell=ws.cell(row=3,column=i)  
    cell.value=  str(policy_number) 
    cell=ws.cell(row=4,column=i)  
    cell.value=  policy_status        
    cell=ws.cell(row=6,column=i)  
    cell.value=  str(policy_effective_date) 
    cell=ws.cell(row=7,column=i)  
    cell.value=  str(policy_expiry_date)
    cell=ws.cell(row=9,column=i)  
    cell.value=  Surname
    cell=ws.cell(row=10,column=i)  
    cell.value= given_name
    cell=ws.cell(row=11,column=i)  
    cell.value= str(birthdate)
    cell=ws.cell(row=12,column=i)  
    cell.value=Gender
    cell=ws.cell(row=13,column=i)  
    cell.value=occupation
    cell=ws.cell(row=16,column=i)  
    cell.value=address_type
    cell=ws.cell(row=17,column=i)  
    cell.value=address
    cell=ws.cell(row=18,column=i)  
    cell.value=city
    cell=ws.cell(row=19,column=i)  
    cell.value=state_province
    cell=ws.cell(row=20,column=i)  
    cell.value=str(postal_code)
    cell=ws.cell(row=21,column=i)  
    cell.value=country
    cell=ws.cell(row=25,column=i)  
    cell.value=product_type
    cell=ws.cell(row=26,column=i)  
    cell.value=product_name
    cell=ws.cell(row=27,column=i)  
    cell.value=str(Face_Amt)
    cell=ws.cell(row=28,column=i)  
    cell.value=str(Premium_Amt)
    cell=ws.cell(row=29,column=i)  
    cell.value=Premium_Frequency
    cell=ws.cell(row=30,column=i)  
    cell.value=First_Unpaid_premium_date
    cell=ws.cell(row=33,column=i)  
    cell.value=str(carrier_name)
    cell=ws.cell(row=34,column=i)  
    cell.value=carrier_code
    cell=ws.cell(row=37,column=i)  
    cell.value=Agency_name
    cell=ws.cell(row=38,column=i)  
    cell.value=Agent_code
    cell=ws.cell(row=39,column=i)  
    cell.value=Commission_Pct
    cell=ws.cell(row=43,column=i)  
    cell.value=under_writing_decision
    cell=ws.cell(row=46,column=i)  
    cell.value=str(Income)
    cell=ws.cell(row=47,column=i)  
    cell.value=calculate_Income_value(value)
    cell=ws.cell(row=48,column=i)  
    cell.value=calculate_Income_score(Income)
    # cell=ws.cell(row=48,column=i)  
    # cell.value=value
    cell=ws.cell(row=50,column=i)  
    cell.value=str(age)
    cell=ws.cell(row=51,column=i)  
    cell.value=calculate_value(value)
    cell=ws.cell(row=52,column=i)  
    cell.value=calculate_score(age)
    cell=ws.cell(row=54,column=i)  
    cell.value=str(smoking_status)
    cell=ws.cell(row=55,column=i)  
    cell.value=calculate_smoking_status_value(value)
    cell=ws.cell(row=56,column=i)  
    cell.value=calculate_smoking_status(smoking_status)
    # cell=ws.cell(row=59,column=i)  
    # cell.value=str(occpation_risk)
    cell=ws.cell(row=60,column=i)  
    cell.value=calculate_occupation_risk_value(occupation)
    cell=ws.cell(row=61,column=i)  
    cell.value=calculate_occupation_risk(occupation)
    cell=ws.cell(row=64,column=i)  
    cell.value=str(family_medical_history)
    cell=ws.cell(row=65,column=i)  
    cell.value=calculate_medical_history_value(value)
    cell=ws.cell(row=66,column=i)  
    cell.value=calculate_medical_history(family_medical_history)
    cell=ws.cell(row=69,column=i)  
    cell.value=height
    cell=ws.cell(row=70,column=i)  
    cell.value=weight
    cell=ws.cell(row=71,column=i)  
    cell.value=calculate_height(height,weight)
    cell=ws.cell(row=72,column=i)  
    cell.value=calculate_height_value(height,weight)
    cell=ws.cell(row=74,column=i)  
    cell.value=str(mvr)
    cell=ws.cell(row=75,column=i)  
    cell.value=calculate_MVR_value(mvr)
    cell=ws.cell(row=76,column=i)  
    cell.value=calculate_MVR(mvr)
    cell=ws.cell(row=79,column=i)  
    cell.value=str(mib)
    cell=ws.cell(row=80,column=i)  
    cell.value=calculate_MIB_value(value)
    cell=ws.cell(row=81,column=i)  
    cell.value=calculate_mib_score(mib)
    cell=ws.cell(row=84,column=i)  
    cell.value=value
    cell=ws.cell(row=85,column=i)  
    cell.value= calculate_credit(value)
    # cell=ws.cell(row=87,column=i)  
    # cell.value=social_media
    # cell=ws.cell(row=88,column=i)  
    # cell.value=value
    # cell=ws.cell(row=89,column=i)
    # cell.value=score
    # cell=ws.cell(row=91,column=i)
    # cell.value=name
    # cell=ws.cell(row=92,column=i)
    # cell.value=description
    # cell=ws.cell(row=95,column=i)
    # cell.value=name
    cell=ws.cell(row=99,column=i)
    cell.value=policy_status_claim(policy_effective_date,First_Unpaid_premium_date,policy_status)
    cell=ws.cell(row=100,column=i)
    cell.value=policy_status_claim_settle(policy_status,random_date)
    cell=ws.cell(row=101,column=i)
    cell.value=policy_status_claim_cause(policy_status)
    
    wb.save('ouput23.xlsx')   